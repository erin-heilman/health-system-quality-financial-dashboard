"""
Excel export functionality for CMS Star Rating Calculator
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List, Dict
import io
from star_rating_calculator import MEASURE_GROUPS


def export_to_excel(results_list: List[Dict], filename: str = None, exec_summary: Dict = None) -> io.BytesIO:
    """
    Export results to Excel with multiple sheets and formatting.

    Args:
        results_list: List of hospital results
        filename: Optional filename to save to disk
        exec_summary: Optional executive summary dictionary (for multi-hospital analysis)

    Returns:
        BytesIO buffer with Excel file
    """
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Use provided executive summary for multi-hospital analysis
    if exec_summary is not None:

        # Executive Summary Sheet 1: System Overview
        ws_overview = wb.create_sheet("Executive_Overview", 0)
        _write_dataframe_to_sheet(ws_overview, exec_summary['System_Overview'], "Executive Overview")

        # Executive Summary Sheet 2: Top Problem Measures (with weighted impact scores)
        ws_problems = wb.create_sheet("Top_Problem_Measures", 1)
        _write_dataframe_to_sheet(ws_problems, exec_summary['Top_Problem_Measures'], "Top Problem Measures (Weighted)")

        # Executive Summary Sheet 3: Group Performance
        ws_groups = wb.create_sheet("Group_Performance", 2)
        _write_dataframe_to_sheet(ws_groups, exec_summary['Group_Performance'], "Group Performance")

        # Executive Summary Sheet 4: Star Rating Opportunities (Quick Wins)
        ws_opps = wb.create_sheet("Star_Rating_Opportunities", 3)
        _write_dataframe_to_sheet(ws_opps, exec_summary['Improvement_Opportunities'], "Star Rating Opportunities")

        # Executive Summary Sheet 5: Action Plan (Weighted Recommendations)
        if not exec_summary['Recommendations'].empty:
            ws_recs = wb.create_sheet("Action_Plan", 4)
            _write_dataframe_to_sheet(ws_recs, exec_summary['Recommendations'], "Action Plan")

    # Original Detailed Sheets
    # Sheet: Summary
    _create_summary_sheet(wb, results_list)

    # Sheet: Group Scores
    _create_group_scores_sheet(wb, results_list)

    # Sheet: Measure Details
    _create_measure_details_sheet(wb, results_list)

    # Sheet: Problem Analysis (Legacy - kept for compatibility)
    _create_problem_analysis_sheet(wb, results_list)

    # Sheet: Methodology
    _create_methodology_sheet(wb)

    # Save to buffer or file
    if filename:
        wb.save(filename)
        print(f"✅ Excel report saved to {filename}")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer


def _create_summary_sheet(wb: Workbook, results_list: List[Dict]):
    """Create summary sheet with key metrics."""
    ws = wb.create_sheet("Summary", 0)

    # Prepare data
    data = []
    for result in results_list:
        row = {
            'Facility_ID': result['provider_id'],
            'Star_Rating': result['star_rating'],
            'Eligible': 'Yes' if result['is_eligible'] else 'No',
            'Peer_Group': result['peer_group'] if result['is_eligible'] else 'N/A',
            'Summary_Score': f"{result['summary_score']:.4f}" if result['summary_score'] else 'N/A',
            'Total_Measures': len(result['measure_details']),
            'Problem_Areas': ', '.join(result['problem_areas']) if result['problem_areas'] else 'None',
            'Strength_Areas': ', '.join(result['strength_areas']) if result['strength_areas'] else 'None'
        }

        # Add group scores
        for group_name in MEASURE_GROUPS.keys():
            score = result['standardized_group_scores'].get(group_name)
            row[f'{group_name}_Score'] = f"{score:.4f}" if score is not None else 'N/A'

        data.append(row)

    df = pd.DataFrame(data)

    # Write to sheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            # Header row formatting
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Star rating color coding
            if r_idx > 1 and df.columns[c_idx-1] == 'Star_Rating' and value:
                if value == 5:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif value == 4:
                    cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
                elif value == 3:
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                elif value == 2:
                    cell.fill = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")
                elif value == 1:
                    cell.fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def _create_group_scores_sheet(wb: Workbook, results_list: List[Dict]):
    """Create detailed group scores sheet."""
    ws = wb.create_sheet("Group_Scores")

    # Prepare data
    data = []
    for result in results_list:
        for group_name in MEASURE_GROUPS.keys():
            std_score = result['standardized_group_scores'].get(group_name)
            raw_score = result['raw_group_scores'].get(group_name)
            weight = result['adjusted_weights'].get(group_name)
            measures = result['measures_per_group'].get(group_name, 0)

            row = {
                'Facility_ID': result['provider_id'],
                'Star_Rating': result['star_rating'],
                'Group': group_name,
                'Standardized_Score': f"{std_score:.4f}" if std_score is not None else 'N/A',
                'Raw_Score': f"{raw_score:.4f}" if raw_score is not None else 'N/A',
                'Weight': f"{weight:.4f}" if weight is not None else 'N/A',
                'Num_Measures': measures,
                'Performance': 'Strong' if std_score and std_score > 0.5 else 'Weak' if std_score and std_score < -0.5 else 'Average'
            }
            data.append(row)

    df = pd.DataFrame(data)

    # Write to sheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Performance color coding
            if r_idx > 1 and df.columns[c_idx-1] == 'Performance':
                if value == 'Strong':
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif value == 'Weak':
                    cell.fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
                elif value == 'Average':
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    ws.freeze_panes = "A2"

    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def _create_measure_details_sheet(wb: Workbook, results_list: List[Dict]):
    """Create measure-level details sheet."""
    ws = wb.create_sheet("Measure_Details")

    # Prepare data
    data = []
    for result in results_list:
        for measure, details in result['measure_details'].items():
            row = {
                'Facility_ID': result['provider_id'],
                'Star_Rating': result['star_rating'],
                'Measure': measure,
                'Original_Value': f"{details['original']:.4f}" if pd.notna(details['original']) else 'N/A',
                'Standardized_Score': f"{details['standardized']:.4f}" if pd.notna(details['standardized']) else 'N/A',
                'vs_National': 'Above' if details['standardized'] > 0 else 'Below' if details['standardized'] < 0 else 'Average',
                'Magnitude': 'Large' if abs(details['standardized']) > 1.5 else 'Moderate' if abs(details['standardized']) > 0.5 else 'Small'
            }
            data.append(row)

    df = pd.DataFrame(data)

    # Write to sheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"

    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def create_executive_summary(results_list: List[Dict]) -> Dict[str, pd.DataFrame]:
    """
    Create comprehensive executive summary for health system leaders.

    Returns:
        Dictionary of DataFrames for different analysis views
    """
    # Filter out hospitals with 'No Data' or 'Insufficient Data' status
    valid_results = [r for r in results_list if isinstance(r.get('star_rating'), int)]

    # 1. SYSTEM-WIDE OVERVIEW
    system_overview = {
        'Total_Hospitals': len(results_list),
        'Hospitals_with_Valid_Ratings': len(valid_results),
        'Hospitals_Insufficient_Data': sum(1 for r in results_list if r.get('star_rating') == 'Insufficient Data'),
        'Hospitals_No_Data': sum(1 for r in results_list if r.get('star_rating') == 'No Data'),
        'Average_Star_Rating': np.mean([r['star_rating'] for r in valid_results]) if valid_results else 0,
        '5_Star_Hospitals': sum(1 for r in valid_results if r.get('star_rating') == 5),
        '4_Star_Hospitals': sum(1 for r in valid_results if r.get('star_rating') == 4),
        '3_Star_Hospitals': sum(1 for r in valid_results if r.get('star_rating') == 3),
        '2_Star_Hospitals': sum(1 for r in valid_results if r.get('star_rating') == 2),
        '1_Star_Hospitals': sum(1 for r in valid_results if r.get('star_rating') == 1),
    }

    # 2. MOST COMMON PROBLEM MEASURES (Top 10)
    problem_measures = {}

    for result in valid_results:
        if 'measure_details' in result:
            for measure, details in result['measure_details'].items():
                if details.get('standardized', 0) < -0.5:  # Significantly below average
                    if measure not in problem_measures:
                        problem_measures[measure] = {
                            'count': 0,
                            'z_scores': [],
                            'hospitals': []
                        }
                    problem_measures[measure]['count'] += 1
                    problem_measures[measure]['z_scores'].append(details.get('standardized', 0))
                    problem_measures[measure]['hospitals'].append(result['provider_id'])

    # Calculate averages and sort by frequency
    problem_measures_data = []
    for measure, data in problem_measures.items():
        avg_z = np.mean(data['z_scores'])
        problem_measures_data.append({
            'Measure': measure,
            'Hospitals_Affected': data['count'],
            'Pct_of_System': f"{(data['count']/len(valid_results)*100) if valid_results else 0:.1f}%",
            'Avg_Z_Score': f"{avg_z:.3f}",
            'Severity': 'CRITICAL' if avg_z < -1.0 else 'HIGH' if avg_z < -0.75 else 'MODERATE',
            'Sample_Hospitals': ', '.join(map(str, data['hospitals'][:3]))
        })

    problem_measures_df = pd.DataFrame(problem_measures_data).sort_values('Hospitals_Affected', ascending=False).head(10)

    # 3. GROUP PERFORMANCE ANALYSIS
    group_performance = []
    for group in ['Mortality', 'Safety of Care', 'Readmission', 'Patient Experience', 'Timely & Effective Care']:
        scores = [r['standardized_group_scores'].get(group) for r in valid_results
                  if 'standardized_group_scores' in r and r['standardized_group_scores'].get(group) is not None]

        if scores:
            hospitals_below = sum(1 for s in scores if s < -0.5)
            avg_score = np.mean(scores)
            group_performance.append({
                'Measure_Group': group,
                'System_Average': f"{avg_score:.3f}",
                'Best_Hospital': f"{max(scores):.3f}",
                'Worst_Hospital': f"{min(scores):.3f}",
                'Hospitals_Below_National': hospitals_below,
                'Pct_Below_National': f"{(hospitals_below/len(scores)*100):.1f}%",
                'Priority': 'HIGH' if avg_score < -0.5 else 'MEDIUM' if avg_score < 0 else 'LOW'
            })

    group_performance_df = pd.DataFrame(group_performance).sort_values('System_Average')

    # 4. STAR RATING IMPROVEMENT OPPORTUNITIES
    from star_rating_calculator import STAR_CUTOFFS
    improvement_opps = []

    for result in valid_results:
        if isinstance(result.get('star_rating'), int) and result.get('peer_group') and result.get('star_rating') < 5:
            current_rating = result['star_rating']
            summary_score = result['summary_score']
            peer_group = result['peer_group']

            # Get cutoff for next star rating
            if peer_group in STAR_CUTOFFS and current_rating < 5:
                next_rating_cutoff = STAR_CUTOFFS[peer_group][current_rating + 1][0]
                gap = next_rating_cutoff - summary_score

                improvement_opps.append({
                    'Facility_ID': result['provider_id'],
                    'Current_Rating': current_rating,
                    'Summary_Score': f"{summary_score:.3f}",
                    'Next_Rating_Cutoff': f"{next_rating_cutoff:.3f}",
                    'Gap_to_Next_Star': f"{gap:.3f}",
                    'Primary_Problem': result.get('problem_areas', ['None'])[0] if result.get('problem_areas') else 'None',
                    'Top_Opportunity': result.get('problem_areas', ['See details'])[0] if result.get('problem_areas') else 'None'
                })

    improvement_opps_df = pd.DataFrame(improvement_opps).sort_values('Gap_to_Next_Star')

    # 5. ACTIONABLE RECOMMENDATIONS
    recommendations = []

    # Based on group performance, create specific recommendations
    if not group_performance_df.empty:
        for _, row in group_performance_df.head(3).iterrows():  # Top 3 problem areas
            group = row['Measure_Group']
            pct = float(row['Pct_Below_National'].rstrip('%'))

            if pct > 50:
                recommendations.append({
                    'Priority': 'CRITICAL',
                    'Focus_Area': group,
                    'Issue': f"{pct:.0f}% of hospitals underperforming in {group}",
                    'Recommendation': f"Implement system-wide quality improvement initiative for {group}. " +
                                     f"Consider centralizing best practices from high-performing hospitals."
                })
            elif pct > 25:
                recommendations.append({
                    'Priority': 'HIGH',
                    'Focus_Area': group,
                    'Issue': f"{pct:.0f}% of hospitals underperforming in {group}",
                    'Recommendation': f"Conduct targeted interventions for struggling hospitals. " +
                                     f"Share best practices from top performers."
                })

    recommendations_df = pd.DataFrame(recommendations) if recommendations else pd.DataFrame()

    return {
        'System_Overview': pd.DataFrame([system_overview]),
        'Top_Problem_Measures': problem_measures_df,
        'Group_Performance': group_performance_df,
        'Improvement_Opportunities': improvement_opps_df,
        'Recommendations': recommendations_df
    }


def _write_dataframe_to_sheet(ws, df: pd.DataFrame, sheet_title: str):
    """Helper function to write a DataFrame to a worksheet with formatting."""
    if df.empty:
        ws.cell(row=1, column=1, value=f"No data available for {sheet_title}")
        return

    # Write to sheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            # Header row formatting
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Priority column color coding
            if r_idx > 1 and 'Priority' in df.columns and df.columns[c_idx-1] == 'Priority':
                if value == 'CRITICAL':
                    cell.fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
                    cell.font = Font(bold=True, color="990000")
                elif value == 'HIGH':
                    cell.fill = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")
                    cell.font = Font(bold=True, color="E69138")
                elif value == 'MEDIUM':
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                elif value == 'LOW':
                    cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

    ws.freeze_panes = "A2"

    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column_letter].width = adjusted_width


def _create_problem_analysis_sheet(wb: Workbook, results_list: List[Dict]):
    """Create problem analysis sheet."""
    ws = wb.create_sheet("Problem_Analysis")

    # Calculate statistics
    problem_counts = {group: 0 for group in MEASURE_GROUPS.keys()}
    strength_counts = {group: 0 for group in MEASURE_GROUPS.keys()}
    avg_scores = {group: [] for group in MEASURE_GROUPS.keys()}

    for result in results_list:
        for group in result['problem_areas']:
            problem_counts[group] += 1

        for group in result['strength_areas']:
            strength_counts[group] += 1

        for group, score in result['standardized_group_scores'].items():
            avg_scores[group].append(score)

    # Prepare summary data
    data = []
    for group in MEASURE_GROUPS.keys():
        scores = avg_scores[group]
        row = {
            'Measure_Group': group,
            'Avg_Score': f"{np.mean(scores):.4f}" if scores else 'N/A',
            'Problem_Count': problem_counts[group],
            'Strength_Count': strength_counts[group],
            'Total_Hospitals': len(scores),
            'Pct_Problem': f"{(problem_counts[group]/len(scores)*100):.1f}%" if scores else 'N/A',
            'Pct_Strength': f"{(strength_counts[group]/len(scores)*100):.1f}%" if scores else 'N/A'
        }
        data.append(row)

    df = pd.DataFrame(data)
    df = df.sort_values('Avg_Score', ascending=True)

    # Write to sheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"

    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def _create_methodology_sheet(wb: Workbook):
    """Create methodology explanation sheet."""
    ws = wb.create_sheet("Methodology")

    methodology_text = [
        ["CMS Hospital Star Rating Methodology", ""],
        ["", ""],
        ["Overview", ""],
        ["The CMS Hospital Star Rating is calculated using a 7-step process:", ""],
        ["", ""],
        ["Step 1: Measure Standardization", ""],
        ["- Each of the 46 quality measures is standardized using Z-scores", ""],
        ["- Z-score = (hospital_value - national_mean) / national_std_dev", ""],
        ["- For 'lower is better' measures, Z-scores are multiplied by -1", ""],
        ["", ""],
        ["Step 2: Group Assignment", ""],
        ["- Measures are assigned to 5 groups:", ""],
        ["  * Mortality (22% weight)", ""],
        ["  * Safety of Care (22% weight)", ""],
        ["  * Readmission (22% weight)", ""],
        ["  * Patient Experience (22% weight)", ""],
        ["  * Timely & Effective Care (12% weight)", ""],
        ["", ""],
        ["Step 3: Calculate Group Scores", ""],
        ["- For each group, calculate the mean of standardized measures", ""],
        ["- Re-standardize group scores across all hospitals", ""],
        ["", ""],
        ["Step 4: Calculate Summary Score", ""],
        ["- Calculate weighted average of group scores", ""],
        ["- Adjust weights if groups are missing (redistribute proportionally)", ""],
        ["", ""],
        ["Step 5: Eligibility Check", ""],
        ["- Must have at least 3 groups with 3+ measures each", ""],
        ["- Must have either Mortality OR Safety of Care group", ""],
        ["", ""],
        ["Step 6: Peer Grouping", ""],
        ["- Hospitals grouped by number of measure groups with 3+ measures (3, 4, or 5)", ""],
        ["", ""],
        ["Step 7: Star Assignment", ""],
        ["- Summary score compared to peer-specific cutoff ranges", ""],
        ["- Stars assigned from 1 (lowest) to 5 (highest)", ""],
        ["", ""],
        ["Data Source", ""],
        [f"CMS Hospital Compare - July 2025 Release", ""],
        ["", ""],
        ["Report Generated By", ""],
        ["CMS Hospital Star Rating Calculator", ""],
        ["https://github.com/yourusername/cms-star-calculator", ""]
    ]

    for r_idx, row in enumerate(methodology_text, 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            # Format headers
            if row[0] and not row[0].startswith(" ") and not row[0].startswith("-") and row[0] != "":
                if r_idx == 1:
                    cell.font = Font(bold=True, size=16, color="366092")
                else:
                    cell.font = Font(bold=True, size=12, color="366092")

    ws.column_dimensions['A'].width = 80
    ws.column_dimensions['B'].width = 40
